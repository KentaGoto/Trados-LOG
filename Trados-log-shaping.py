import bs4
import openpyxl
import shutil
import sys


# エクセルへの入力
def inputExcel(countStr, count, ws, s):
    if count == 1:
        a = 'A' + countStr
        ws[a].value = s
    elif count == 2:
        b = 'B' + countStr
        ws[b].value = int(s)
    elif count == 3:
        c = 'C' + countStr
        ws[c].value = int(s)
    elif count == 4:
        d = 'D' + countStr
        ws[d].value = int(s)
    elif count == 5:
        # 処理対象のHTMLでCross File Repに空文字がある場合に対応
        if type(s) == str:
            if s == "":
                s_zero = "0"
                print('crossfileRep -> ' + s_zero)
                e = 'E' + countStr
                ws[e].value = int(s_zero)
            else:
                print('crossfileRep -> ' + s)
                e = 'E' + countStr
                ws[e].value = int(s)
    elif count == 6:
        f = 'F' + countStr
        ws[f].value = int(s)
    elif count == 7:
        g = 'G' + countStr
        ws[g].value = int(s)
    elif count == 8:
        h = 'H' + countStr
        ws[h].value = int(s)
    elif count == 9:
        i = 'I' + countStr
        ws[i].value = int(s)
    elif count == 10:
        j = 'J' + countStr
        ws[j].value = int(s)
    elif count == 11:
        k = 'K' + countStr
        ws[k].value = int(s)
    elif count == 12:
        l = 'L' + countStr
        ws[l].value = int(s)
    elif count == 13:
        m = 'M' + countStr
        ws[m].value = int(s)


if __name__ == '__main__':
    s = sys.argv[1]
    s_base = sys.argv[2]

    log_file = s.strip('\"')

    # パーサー
    soup = bs4.BeautifulSoup(open(log_file, 'r', encoding="utf-16"), 'html.parser')
    trs = soup.find_all('tr')

    # 結果が入るエクセルの準備
    xlsxTemplate = "log.xlsx"
    results_dir = './results'
    resultsFile = results_dir + '/' + s_base + '_' + xlsxTemplate
    shutil.copyfile(xlsxTemplate, resultsFile)

    # Excelを準備
    wb = openpyxl.load_workbook(resultsFile)
    ws = wb['Characters']
    ws2 = wb['Words']

    # Main
    countT = 1
    flag = 0
    count = 1
    for tr in trs:
        for td in tr.find_all('td'):
            s = td.text
            
            # ワードカウント分が出てきたら新たにセットする
            if s == 'Total (words)':
                flag += 1
                countT = 2
                count = 1

            countStr = str(countT)

            if flag == 0:
                inputExcel(countStr, count, ws, s) # Charactorsシートに入力する
            elif flag <= 1:
                inputExcel(countStr, count, ws2, s) # Wordsシートに入力する

            if count <= 12:
                count += 1
            elif count == 13:
                count = 1

        countT += 1 # エクセルの入力行
    
    # Excelを閉じて保存
    wb.close()
    wb.save(resultsFile)
